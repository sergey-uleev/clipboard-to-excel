# Как записывать данные в Ехcel
import datetime  # Библиотека для работы с датой и временем
import openpyxl
from utils import is_next_day

class Excel:
    filename = ''
    book = None
    sheet = None
    def __init__(self, _filename):
        self.filename = _filename
        self.__open(self.filename)

    def __open(self, filename):
        self.book = openpyxl.load_workbook(filename)
        self.sheet = self.book.active

    def write(self, text):
        row = search_empty(self.sheet)
        rec_num = 1
        if row > 1:
            last_date = datetime.datetime.strptime(self.__get(row - 1, 3), '%d-%m-%Y %H:%M:%S').date()
            now = datetime.datetime.now()
            if not is_next_day(last_date, now):
                rec_num = self.__get(row - 1, 1) + 1
        self.__write_cell(row, 1, rec_num)
        self.__write_cell(row, 2, text)
        time = datetime.datetime.now().strftime('%d-%m-%Y %H:%M:%S')
        self.__write_cell(row, 3, time)
        self.__save_and_close()

    def __write_cell(self, row, col, text):
        self.sheet.cell(row=row, column=col).value = text

    def __save_and_close(self):
        self.book.save(self.filename)
        self.book.close()
    
    def __get(self, row, col):
        return self.sheet.cell(row = row, column = col).value

def search_empty(sheet):
    empty = 1
    while sheet.cell(row = empty, column = 2).value != None:
        empty += 1
    return empty
